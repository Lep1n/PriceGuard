import os
import json
import logging
import subprocess
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Directory Configuration
LOG_DIR = "logs"
SCREENSHOT_DIR = "screenshots"
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(SCREENSHOT_DIR, exist_ok=True)

# File Logging Setup
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(os.path.join(LOG_DIR, "app.log"), encoding="utf-8"),
        logging.StreamHandler()
    ]
)

SAVED_KW_FILE = "last_keywords.json"

def load_config(config_path="config/config.json"):
    """Loads runtime configurations from JSON file."""
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logging.error(f"Error reading config: {e}")
    return {
        "app_version": "v0.2.2",
        "currency": "USD",
        "headless": False,
        "browser_timeout_ms": 15000,
        "output_file": "output.xlsx"
    }

def export_results(records, output_filepath="output.xlsx", export_format="Excel (.xlsx)"):
    """Exports dataset to Excel (.xlsx), CSV, or JSON based on user selection."""
    if not records:
        logging.warning("[EXPORT] No records provided for export.")
        return output_filepath

    df = pd.DataFrame(records)

    if "CSV" in export_format:
        target_path = "output.csv"
        df.to_csv(target_path, index=False, encoding="utf-8-sig")
        logging.info(f"[EXPORT] CSV dataset saved successfully to: '{target_path}'")
        return target_path
    elif "JSON" in export_format:
        target_path = "output.json"
        df.to_json(target_path, orient="records", indent=4, ensure_ascii=False)
        logging.info(f"[EXPORT] JSON dataset saved successfully to: '{target_path}'")
        return target_path
    else:
        export_styled_excel(records, output_filepath)
        return output_filepath

def export_styled_excel(records, output_filepath="output.xlsx"):
    """Exports records to Excel with executive openpyxl styling."""
    df = pd.DataFrame(records)
    df.to_excel(output_filepath, index=False)

    wb = load_workbook(output_filepath)
    ws = wb.active
    ws.title = "PriceGuard_Report"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
    )

    for col_num, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num in range(2, len(df) + 2):
        for col_num, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.font = regular_font

            if col_name in ["Price_USD"]:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = bold_font

            elif col_name in ["Original_Price", "Currency"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = bold_font

            elif col_name == "Stock_Status":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                val_str = str(cell.value).upper()
                if "IN STOCK" in val_str:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="27AE60")
                elif "OUT OF STOCK" in val_str:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="C0392B")
                else:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="E67E22")

            elif col_name == "Extraction_Status":
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 6, 14), 85)

    ws.row_dimensions[1].height = 24
    wb.save(output_filepath)
    logging.info(f"[EXPORT] Executive-styled Excel saved: '{output_filepath}'")

def open_file_or_folder(filepath="output.xlsx"):
    """Safely opens output file or opens parent folder in Windows Explorer."""
    abs_path = os.path.abspath(filepath)
    if not os.path.exists(abs_path):
        from tkinter import messagebox
        messagebox.showerror("Error", f"File '{filepath}' does not exist yet. Run audit first!")
        return

    try:
        os.startfile(abs_path)
    except Exception:
        subprocess.Popen(f'explorer /select,"{abs_path}"')

def open_screenshots_folder():
    """Opens screenshots directory directly in Explorer."""
    abs_path = os.path.abspath(SCREENSHOT_DIR)
    os.makedirs(abs_path, exist_ok=True)
    try:
        os.startfile(abs_path)
    except Exception:
        subprocess.Popen(f'explorer "{abs_path}"')